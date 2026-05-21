import argparse
import csv
from pathlib import Path

import openpyxl


EXPECTED_HEADER = ["행정동코드", "시도명", "시군구명", "읍면동명", "생성일자", "말소일자"]


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def administrative_city_code(code):
    return f"{code[:2]}00000000"


def administrative_district_code(code, district_name):
    if not district_name:
        return administrative_city_code(code)
    return f"{code[:5]}00000"


def convert(source_path, output_path):
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [normalize(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
        raise ValueError(f"Unexpected header: {header}")

    rows = []
    seen_dong_codes = set()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        code = normalize(row[0])
        city_name = normalize(row[1])
        district_name = normalize(row[2])
        dong_name = normalize(row[3])
        deleted_at = normalize(row[5])

        if not code or deleted_at or not dong_name:
            continue
        if code in seen_dong_codes:
            continue

        seen_dong_codes.add(code)
        city_code = administrative_city_code(code)
        district_code = administrative_district_code(code, district_name)
        rows.append([
            city_code,
            city_name,
            district_code,
            district_name or city_name,
            code,
            dong_name,
        ])

    rows.sort(key=lambda item: (item[0], item[2], item[4]))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)
        writer.writerow([
            "city_code",
            "city_name_ko",
            "district_code",
            "district_name_ko",
            "dong_code",
            "dong_name_ko",
        ])
        writer.writerows(rows)

    return len(rows)


def main():
    parser = argparse.ArgumentParser(
        description="Convert Korean administrative dong workbook to Honeytong region seed CSV."
    )
    parser.add_argument("source", type=Path, help="KIKcd_H xlsx source path")
    parser.add_argument("output", type=Path, help="Honeytong region seed CSV output path")
    args = parser.parse_args()

    count = convert(args.source, args.output)
    print(f"Converted {count} administrative dong rows to {args.output}")


if __name__ == "__main__":
    main()
